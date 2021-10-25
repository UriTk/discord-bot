# Utility file which holds external classes and functions, planning on seperating this to seperate files later
import discord
from discord.ext import menus
from typing import List, Dict
from openpyxl import load_workbook
from dataclasses import dataclass
from UtilityFiles.constants import COLOR, RARITY_MAP, NO_IMAGE_URL, ROLE, RARITY_VALUE, CONVERT
import json


class NoGachas(Exception):
    pass


# made by the old bot dev not sure how it works but it returns a generator
# (There was actually an older older older bot dev. It was the first version of this Gacha code.)
# (My predecessor dev re-wrote most of his code. I believe this is some of the only stuff he kept from the old code.)
def exel_generator(filename):
    """Opens and loads an xlsx file into a generator"""
    book = load_workbook(filename)
    sheet = book.active
    rows = sheet.max_row
    cols = sheet.max_column

    def item(i, j):
        return (sheet.cell(row=3, column=j).value, sheet.cell(row=i, column=j).value)

    return (dict(item(i, j) for j in range(2, cols + 1)) for i in range(4, rows + 1))


# used the old function to create a new one that returns a dict of every gacha
def exel_reader(filename):
    result = {}
    for item in exel_generator(filename):
        id = item.pop("ID")
        result[str(id)] = item
    return result


#  Temporarily putting this global here, for Achievements.
GACHA_DATA = exel_reader("UtilityFiles/GACHAUNITS.xlsx")


def dump_data(data, file):
    """Dumps a dictionary into a json file"""
    file.seek(0)
    file.truncate(0)
    json.dump(data, file, indent=4)


def create_cache(user_id):
    """Creates a cache for a user if they have never been stored, returns an empty inventory"""
    empty_inventory = {"coins": 0, "tickets": 0, "special_tickets": 0, "super_tickets": 0, "ultra_tickets": 0,
                       "gachas": {}, "purchases": 0, "rolls": 0, "conversions": 0, "consecutive_logins": 0,
                       "last_login": 0,
                       "achievements": {"1":0, "2":0, "3":0, "4":0, "5":0, "6":0, "7":0, "8":0, "9":0, "10":0, "11":0,
                                        "12":0, "13":0, "14":0, "15":0, "16":0, "17":0, "18":0, "19":0, "20":0, "21":0,
                                        "22":0, "23":0, "24":0, "25":0, "26":0, "27":0, "28":0, "29":0, "30":0, "31":0,
                                        "32":0, "33":0, "34":0, "35":0, "36":0, "37":0, "38":0, "39":0, "40":0, "41":0,
                                        "42":0, "43":0, "44":0, "45":0, "46":0, "47":0, "48":0, "49":0, "50":0, "51":0,
                                        "52":0, "53":0, "54":0, "55":0, "56":0, "57":0, "58":0, "59":0, "60":0, "61":0,
                                        "62":0, "63":0, "64":0, "65":0, "66":0, "67":0}}
    # I hate this, but it's better to do this than to make sure the achievement exists in "achievements" every single
    # time in the achievements checks. Find a better alternative ASAP
    with open("UtilityFiles/user_info.json", "r+") as f:
        stored_info = json.load(f)
        stored_info[str(user_id)] = empty_inventory
        dump_data(data=stored_info, file=f)

    return empty_inventory


def get_inventory(user_id):
    """Gets the inventory of a user from a user info dictionary"""
    with open("UtilityFiles/user_info.json", "r") as f:
        stored_info = json.load(f)
    try:  # tries getting the inventory from the file
        inventory = stored_info[str(user_id)]
    except KeyError:  # creates a cache of the user if they have never been in the file
        inventory = create_cache(user_id)
    return inventory


def get_score(user, excel_data):
    rarity_value = {}
    for setting in RARITY_VALUE:  # We grab the settings for how often a rarity should be put on the pool.
        rarity_value[setting.split(":")[0]] = int(setting.split(":")[1])
    try:
        cards = get_gachas(user, excel_data)
    except NoGachas:
        cards = None
    totalvalue = 0
    if cards is not None:
        for card in cards:
            value = rarity_value[card.rarity]  # Reads it's rarity.
            while card.amount:
                totalvalue += value
                card.amount -= 1
    inv = get_inventory(user.id)
    totalvalue += inv["coins"]

    return totalvalue



def achievement_checker(user, ach_info, value):
    # This is where the main checks for achievements will happen. Will be called a lot.
    # Probably a good idea to separate sections of code on categories, for organization and functionality's sake.
    # This is going to suck, no way to avoid it. Separate into its own file ASAP? Organization's sake.

    # <editor-fold desc="Idea... Proposal? Possibility:">
    # Coding achievements in a traditional videogame sense:
    #
    # Instead of making the code check if a certain quota is met using the user's global statistics, i.e.
    # ###----------------------Assuming we work based on the user's total gachas-----------------###
    #   for unit in usergachas:  (Get total Gachas)
    #       total_unit_amount += int(unit[1])
    #       (Actual absolute literal shit ton of checks to see 'if' there's X amount of characters from X series)
    #   (More 'if' checks for stuff like Hoarder)
    # #--------------------------------------------------------------------------------------------#
    # We could instead start adding on to a user-specific achievement-independent tracking.
    # For instance, in the current example of "UNIT" achievements:
    #
    # Let's assume the character rolled was Goku. Getting him would give the user progress towards Hoarder, 90's kid
    # and possibly some other achievements. We'd make that check:
    # ###----------------------Assuming we work based on the user's latest-rolled gacha----------###
    #   (lot of "if" checks to see if he's from any specific series, for stuff like the "Magical Girl" achievement.)
    #
    #    #  90's Kid
    #   if (ach_value.show == "Dragon Ball Z" | "Yu Yu Hakusho" | "Sailor Moon" | "Cardcaptor Sakura" |
    #   "Pokemon" | "Revolutionary Girl Utena" | "Neon Genesis Evangelion" | "Cowboy Bebop"):
    #       user_inventory["achievements"]["10"] += 1
    # #--------------------------------------------------------------------------------------------#
    # Assuming Goku is their only owned Gacha, their entry on user_info.json would look a little bit like:
    #   "achievements": {
    #   "1":1,
    #   "10":1
    #   }
    # How much you need for you do actually "Have" each achievement would be on achievement.json
    # Note: Said values added to achievements.json. "goals" with a value of 1 should be treated as bools, either 0 or 1.
    # I'm pretty sure making them actual bools would conflict with user_info.json
    #
    # Not only would this be a bit more optimal, since we don't have to do any unnecessary checks in a "for" loop,
    # but this would also let us tell the user how far into any non-secret achievement they are.
    # We can check if they've met the quota for any unowned achievement at the end of the code.
    #
    # The only clear downsides are:
    # -This needs a one-time "catch-up" code, so we can update all the current users in
    # user_info with the right achievement stats for what they currently have. Otherwise, a person with 500 gachas
    # might have a progress of 3/250 in "Hoarder"
    # -Necessary to keep in mind the removal of units via trades & conversions.
    # Keep in mind, if someone has 5 Narutos, and converts or trades one away, they should still keep their achievement.
    #
    #
    # After heavy consideration, this is... Mostly good, I think? It has a small few flaws. Worth doing it and
    # going around the flaws, handling the specific achievements differently. Current found flaws:
    #
    # -The achievements:
    # "I'll Be Here All Week", "You're Going to Be a Star!", "This is How I Win" & "Frugal", "Show-off"
    # don't have 'additive' goals by nature.
    # Frugal needs you to own 100 coins at once, for instance. Can be handled separately, no problem.
    # Get them with specific "ach_info" values called wherever necessary, i.e. the !daily command.
    #
    # -"Fan" & "Big Fan needs us to check all gachas anyway.
    # Is this still more optimal even with them included? Should be.
    # Should check if user has these achievements before doing their checks.
    #
    # -"Zoologist" could possibly need a check of all the user owned gachas. This is probably fine, as we can just do
    # the check when the drawn unit is a cat, dog or any other animal.
    # The same goes for the achievements "Chun III", "New Horizons", "Gentleman Thief", "Nice." & "NICE."
    # </editor-fold>

    # Temp.TODO:Move this to somewhere permanent like constants.py before pushing, we don't need to call this everytime.
    with open("UtilityFiles/achievements.json", "r") as f:
        achievements = json.load(f)  # list of id #'s to make a pool


    user_inventory = get_inventory(str(user.id))
    user_achievements = user_inventory["achievements"]
    # Note: We always check if the user has the achievement before the achievement's actual checks for readability
    # This isn't always the best, however. Optimization-wise.
    # Consider making the achievement's check first for the harder achievements.
    # i.e. Cat person, Does Not Compute, New Horizons
    # "UNITADD" achievements. Achievements that are called if a unit is added to the user's inventory. Roll & Store.
    # Currently all unit-related achievements are here. That needs change.
    if ach_info == "unitadd":

        # 1: Hoarder & 56: Weeb
        if user_achievements["56"] < 500:
            user_achievements["56"] += 1
            if user_achievements["1"] < 250:
                user_achievements["1"] += 1

        # 2: Fan & 31 : Big Fan. (Gross, fix this, Uri. And make it not shit. And faster.)
        if user_achievements["31"] == 0:
            showdict = {}
            usergachas = get_gachas(user, excel_data=GACHA_DATA)
            if user_achievements["2"] == 0:
                for gacha in usergachas:
                    if gacha.show in showdict.keys():  # I'd rather not have to do this check. Possible solutions?
                        showdict[gacha.show] += gacha.amount
                        if showdict[gacha.show] >= 5:
                            user_achievements["2"] = 1
                    else:
                        showdict[gacha.show] = gacha.amount
            # Gross. Same loop, except if we need "Big Fan" and not "Fan", we check for 10 instead.
            else:  # Is the small optimization worth the grossness?
                for gacha in usergachas:
                    if gacha.show in showdict.keys():
                        showdict[gacha.show] += gacha.amount
                        if showdict[gacha.show] >= 25:
                            user_achievements["31"] = 1
                    else:
                        showdict[gacha.show] = gacha.amount

        # 3: Collector & 32: Connossieur
        print(value.rarity)
        if value.rarity == "UR":
            if user_achievements["32"] < 25:
                user_achievements["32"] += 1
                if user_achievements["3"] < 10:
                    user_achievements["3"] += 1

        # 10: 90s Kid
        if user_achievements["10"] < 25:
            if (value.show in ("Dragon Ball Z", "Yu Yu Hakusho", "Sailor Moon", "Cardcaptor Sakura",
               "Pokemon", "Revolutionary Girl Utena", "Neon Genesis Evangelion", "Cowboy Bebop")):
                user_achievements["10"] += 1

        # 11: Magical Girl
        if user_achievements["11"] < 15:
            if value.show in ("Sailor Moon", "Cardcaptor Sakura", "Revolutionary Girl Utena"):
                user_achievements["11"] += 1

        # 12: Fujoshi
        if user_achievements["12"] < 15:
            if value.show in ("Yuri!!! On Ice", "Haikyuu", "Banana Fish", "Fruits Basket"):
                user_achievements["12"] +=1

        # 13: Your 30-something Older Brother
        if user_achievements["13"] < 15:
            if value.show in ("Cowboy Bebop", "Berserk", "Rurouni Kenshin", "Akira"):
                user_achievements["13"] += 1

        # 14: Shonen
        if user_achievements["14"] < 25:
            if value.show in ("Naruto", "One Piece", "Hunter x Hunter", "Bleach", "Demon Slayer", "My Hero Academia"):
                user_achievements["14"] += 1

        # 17: Believe It!
        if user_achievements["17"] < 5:
            if value.card_id == 1:
                user_achievements["17"] += 1

        # 19: Gamer
        if user_achievements["19"] < 30:
            if value.header == "Video Game":
                user_achievements["19"] += 1

        # 20: Zoologist. This is readable, but it sucks. Only a couple of rarer units trigger this, though, so it's fine
        if user_achievements["20"] == 0:
            if value.descriptor in ("Dog", "Cat", "Animal"):
                usergachas = get_gachas(user, excel_data=GACHA_DATA)
                cat = False
                dog = False
                animal = False
                for gacha in usergachas:
                    if gacha.descriptor == cat:
                        cat = True
                    elif gacha.descriptor == dog:
                        dog = True
                    elif gacha.descriptor == animal:
                        animal = True
                if cat and dog and animal:
                    user_achievements["20"] = 0

        # 28: Nice.
        if user_achievements["28"] == 0:
            if value.card_id == 69 | 420:
                if set({'69', '420'}).issubset(user_inventory["gachas"].keys()):
                    user_achievements["28"] = 1

        # 39: Ghibli Fanboy
        if user_achievements["39"] < 30:
            if value.ghibli is not None:
                user_achievements["39"] += 1

        # 41: Chun III
        if user_achievements["41"] == 0:
            if value.card_id == 464 | 465 | 477:
                if set({'464', '465', '477'}).issubset(user_inventory["gachas"].keys()):
                    user_achievements["41"] = 1

        # 42: Cat Person
        if user_achievements["42"] < 5:
            if value.descriptor == "Cat":
                user_achievements["42"] += 1

        # 43: Does Not Compute
        if user_achievements["43"] < 5:
            if value.descriptor == "Robot":
                user_achievements["43"] += 1

        # 44: New Horizons
        if user_achievements["44"] == 0:
            if 491 <= value.card_id <= 500:
                if set({range(491, 500)}).issubset(user_inventory["gachas"].keys()):
                    user_achievements["41"] = 1

        # 47: Pokemon Trainer
        if user_achievements["47"] < 6:
            if value.descriptor == "Pokemon":
                user_achievements["47"] += 1

        # 49: Monopoly. Too tired to do this one well at the moment.

        # 51: NICE.
        if user_achievements["51"] == 0:
            if value.card_id == 69 | 169 | 269 | 369 | 469:
                if set({'69', '169', '269', '369', '469'}).issubset(user_inventory["gachas"].keys()):
                    user_achievements["28"] = 1

        # 64: Gentleman Thief. The achievement says "Own all 4 Lupin units of a different color suit"
        # However, my gacha sheet only has 3 Lupins. Verify before coding.


    return


def is_allowed(ctx):  # A role/owner check for some commands

    groups = ctx.message.author.roles  # The user's roles
    admin = False
    for group in groups:  # Checks if any of the
        if group.permissions.manage_guild:
            admin = True
        else:
            if str(group.name).lower() == ROLE:
                admin = True

    if str(ctx.message.guild.owner_id) == str(ctx.message.author.id):
        admin = True
    if admin:
        return True
    return False


def add_card(user: discord.Member, card_id: str):
    """Adds a card to a user's inventory"""
    with open("UtilityFiles/user_info.json", "r+") as f:
        stored_info = json.load(f)
        if card_id in stored_info[str(user.id)]["gachas"].keys():
            stored_info[str(user.id)]["gachas"][card_id] += 1
        else:
            stored_info[str(user.id)]["gachas"][card_id] = 1
        dump_data(data=stored_info, file=f)


def remove_card(user: discord.Member, card_id: str):
    """Removes a card to a user's inventory"""
    with open("UtilityFiles/user_info.json", "r+") as f:
        stored_info = json.load(f)
        if int(stored_info[str(user.id)]["gachas"][card_id]) > 1:
            stored_info[str(user.id)]["gachas"][card_id] -= 1
        else:
            stored_info[str(user.id)]["gachas"].pop(card_id)
        dump_data(data=stored_info, file=f)


@dataclass
class GachaCard:
    card_id: int
    name: str
    rarity: str
    show: str
    header: str
    descriptor: str
    ghibli: str
    image_url: str
    amount: int
    stats: Dict[str, int]

    def generate_embed(self, sort_type=None, show_footer=True):
        """Generates an embed with basic info about a card"""
        card_embed = discord.Embed(
            title=f"#**{self.card_id} | {self.name}**",
            colour=COLOR
        )
        card_embed.set_image(url=self.image_url)
        card_embed.add_field(name="**Rarity:**", value=RARITY_MAP[self.rarity])
        card_embed.add_field(name="**Show:**", value=self.show)
        if show_footer:
            if sort_type is None:
                card_embed.set_footer(text=f"Amount: {self.amount}")
            else:
                card_embed.set_footer(text=f"Amount: {self.amount}\nSort: {sort_type}")

        return card_embed

    def generate_stats_embed(self):
        """Generates an embed with stats info for a card"""
        stats_embed = discord.Embed(
            title=f"#**{self.card_id} | {self.name}**",
            colour=COLOR)
        stats_embed.set_image(url=self.image_url)
        for name, stat in self.stats.items():
            stats_embed.add_field(name=f"**{name.lower().title()}:**", value=stat, inline=False)

        return stats_embed


def get_gachas(user: discord.Member, excel_data) -> List[GachaCard]:
    gachas = get_inventory(user.id)["gachas"]
    if gachas:  # if the result isn't empty
        cards = []
        for gacha_id in gachas.keys():
            gacha_info = excel_data[gacha_id]
            try:
                card_stats = {"height": int(gacha_info["Height"]),
                              "intelligence": int(gacha_info["Intelligence"]),
                              "speed": int(gacha_info["Speed"]),
                              "power": int(gacha_info["Power"])}
            except TypeError:
                card_stats = {"height": 0,
                              "intelligence": 0,
                              "speed": 0,
                              "power": 0}
            card_stats["total"] = sum(card_stats.values())

            card = GachaCard(
                card_id=int(gacha_id),
                name=gacha_info["CHARACTER"],
                rarity=gacha_info["Grade"],
                show=gacha_info["Show"],
                header=gacha_info["Subgroup"],
                descriptor=gacha_info["Descriptor"],
                ghibli=gacha_info["Ghibli?"],  # This one is pretty unnecessary, but it's on the excel, so why not
                image_url=gacha_info["IMAGE"] if gacha_info["IMAGE"] is not None else NO_IMAGE_URL,
                amount=int(gachas[gacha_id]),
                stats=card_stats)
            cards.append(card)

        return cards
    else:
        raise NoGachas("The user has no Gachas in their inventory")


def get_pool(rarity) -> List[GachaCard]:
    """Returns the Gacha Pool"""
    gachas = exel_reader("UtilityFiles/GACHAUNITS.xlsx")
    if rarity == "tickets":
        with open("UtilityFiles/gacha_pool.json", "r") as f:
            pool_ids = json.load(f)  # list of id #'s to make a pool
    elif rarity == "common":  # Note: Common and rare pools aren't actually made for any specific rolls, so they don't have tickets.
        with open("UtilityFiles/gacha_pool_common.json", "r") as f:
            pool_ids = json.load(f)
    elif rarity == "rare":
        with open("UtilityFiles/gacha_pool_rare.json", "r") as f:
            pool_ids = json.load(f)
    elif rarity == "special_tickets":
        with open("UtilityFiles/gacha_pool_special.json", "r") as f:
            pool_ids = json.load(f)
    elif rarity == "super_tickets":
        with open("UtilityFiles/gacha_pool_super.json", "r") as f:
            pool_ids = json.load(f)
    elif rarity == "ultra_tickets":
        with open("UtilityFiles/gacha_pool_ultra.json", "r") as f:
            pool_ids = json.load(f)
    elif rarity == "all":
        pool_ids = range(1, len(gachas) + 1)

    pool = []
    for pool_id in pool_ids:
        gacha_info = gachas[str(pool_id)]  # retrieves gacha with correct id
        if gacha_info["Subgroup"] is None:
            gacha_info["Subgroup"] = "Normal"
        try:
            card_stats = {"height": int(gacha_info["Height"]),
                          "intelligence": int(gacha_info["Intelligence"]),
                          "speed": int(gacha_info["Speed"]),
                          "power": int(gacha_info["Power"])}
        except TypeError:
            card_stats = {"height": 0,
                          "intelligence": 0,
                          "speed": 0,
                          "power": 0}
        card_stats["total"] = sum(card_stats.values())
        card = GachaCard(
            card_id=int(pool_id),
            name=gacha_info["CHARACTER"],
            rarity=gacha_info["Grade"],
            show=gacha_info["Show"],
            header=gacha_info["Subgroup"],
            descriptor=gacha_info["Descriptor"],
            ghibli=gacha_info["Ghibli?"],  # This one is pretty unnecessary, but it's on the excel, so why noty
            image_url=gacha_info["IMAGE"] if gacha_info["IMAGE"] is not None else NO_IMAGE_URL,
            amount=None,
            stats=card_stats)
        pool.append(card)
    return pool


def is_stat_embed(embed: discord.Embed):
    return embed.fields[0].name == "**Height:**"


def id_from_embed(embed: discord):
    return int(embed.title[3:].split("|")[0].strip())


class CardMenu(menus.MenuPages):
    """Shows the collection as a menu of cards"""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.current_sort = 0  # creates a current sort variable to sort through them

    @menus.button('🔀')
    async def on_shuffle(self, payload):
        self.current_sort += 1
        if self.current_sort >= len(SORT_LIST):  # loops back to 0 when finished going through sorts
            self.current_sort = 0

        cards = self._source.cards
        await self.change_source(CardMenuSource(cards, self.current_sort))

    @menus.button('📊', position=menus.Last(2))
    async def show_stats(self, payload):
        cards = self._source.cards
        card_id = id_from_embed(self.message.embeds[0])  # parses the embed and fetches the card id
        current_card = [card for card in cards if card.card_id == card_id][0]

        if is_stat_embed(self.message.embeds[0]):  # if the embed is a stats embed
            await self.message.edit(embed=current_card.generate_embed())
        else:  # if the embed is a regular embed
            await self.message.edit(embed=current_card.generate_stats_embed())

    # necessary to override the default black square in discord.Menus
    @menus.button('\N{BLACK SQUARE FOR STOP}\ufe0f', position=menus.Last(3))
    async def stop_pages(self, payload):
        """stops the pagination session."""
        self.stop()


class CardMenuSource(menus.ListPageSource):
    """Source for CardMenu"""

    def __init__(self, cards: GachaCard, current_sort: int = 0):
        self.current_sort = current_sort
        self.cards = SORT_LIST[self.current_sort](cards)  # calls the current sort func on the cards list
        super().__init__([*range(len(cards))], per_page=1)

    async def format_page(self, menu, page):
        return self.cards[page].generate_embed(SORT_LIST[self.current_sort].__doc__)


class ShowGachaMenu(menus.Menu):
    """Creates a menu for the show command"""

    def __init__(self, card, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.card = card  # creates a current sort variable to sort through them

    async def send_initial_message(self, ctx, channel):
        return await channel.send(embed=self.card.generate_embed())

    @menus.button('📊', position=menus.Last(2))
    async def show_stats(self, payload):
        if is_stat_embed(self.message.embeds[0]):  # if the embed is a stats embed
            await self.message.edit(embed=self.card.generate_embed())
        # if the embed is a regular embed
        else:
            await self.message.edit(embed=self.card.generate_stats_embed())


class LotteryGachaMenu(menus.Menu):
    """Creates a menu for the show command"""

    def __init__(self, card, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.card = card  # creates a current sort variable to sort through them

    async def send_initial_message(self, ctx, channel):
        return await channel.send(embed=self.card.generate_embed(None, False))


class ShopConfirmMenu(menus.Menu):
    """Creates a menu for the show command"""

    def __init__(self, ctx, item, is_gacha, *args, **kwargs):
        super().__init__(timeout=30.0, delete_message_after=True, *args, **kwargs)
        self.ctx = ctx
        self.item = item
        self.is_gacha = is_gacha
        self.confirmation = None

    async def send_initial_message(self, ctx, channel):
        if self.is_gacha:
            self.confirmation = await channel.send(
                embed=discord.Embed(title=f"You are about to buy {self.item.name}. React with ✔ to confirm.",
                                    colour=COLOR))
            return await channel.send(embed=self.item.generate_embed())
        else:
            return await channel.send(
                embed=discord.Embed(title=f"You are about to buy a {self.item}. React with ✔ to confirm.",
                                    colour=COLOR))

    @menus.button(emoji='📊')
    async def show_stats(self, payload):
        if self.is_gacha:
            if is_stat_embed(self.message.embeds[0]):  # if the embed is a stats embed
                await self.message.edit(embed=self.item.generate_embed())
            # if the embed is a regular embed
            else:
                await self.message.edit(embed=self.item.generate_stats_embed())

    @menus.button('❌', position=menus.Last(2))
    async def cancel(self, payload):
        await self.message.delete()
        if self.is_gacha:
            await self.confirmation.delete()
        return await self.message.channel.send(
            embed=discord.Embed(title=f"Purchase canceled.", colour=COLOR))
        self.stop()

    @menus.button('✔️', position=menus.Last(3))
    async def confirm(self, payload):
        inventory = get_inventory(self.ctx.author.id)
        if self.is_gacha:
            gachas = inventory["gachas"]  # retrieves the user's gachas
            add_card(self.ctx.author, str(self.item.card_id))  # adds a new card to the gachas

            # finds out how many of the card the user now has
            if str(self.item.card_id) in gachas.keys():  # if the user has the card already
                self.item.amount = gachas[
                                       str(
                                           self.item.card_id)] + 1  # card.amount is set to how many of the card the user now has
                message = f"Congratulations, you just got another {self.item.name}! You now have {self.item.amount} {self.item.name}'s"
            else:  # the card must be new
                self.item.amount = 1  # the amount is by default 1
                message = f"Congratulations, you just got a new card, {self.item.name}!"

            if self.item:  # if there is a card with the correct id
                await self.message.delete()
                await self.confirmation.delete()
                await self.ctx.send(embed=discord.Embed(title=message, colour=COLOR))
                await ShowGachaMenu(self.item).start(self.ctx)  # creates a menu to show the card
        else:
            # add card, remove tickets here
            await self.message.delete()
            await self.message.channel.send(embed=discord.Embed(
                title=f"You purchased a {self.item}.",
                colour=COLOR))
            if self.item == "Ticket":
                self.item = "tickets"
            elif self.item == "Special Ticket":
                self.item = "special_tickets"
            elif self.item == "Super Ticket":
                self.item = "super_tickets"
            elif self.item == "Ultra Ticket":
                self.item = "ultra_tickets"
        with open("UtilityFiles/user_info.json", "r+") as f:
            stored_user_info = json.load(f)
            if not self.is_gacha:
                stored_user_info[str(self.ctx.author.id)][self.item] += 1
            stored_user_info[str(self.ctx.author.id)]["coins"] = inventory["coins"]
            stored_user_info[str(self.ctx.author.id)]["purchases"] += 1
            dump_data(data=stored_user_info, file=f)
        self.stop()


class ConvertGachaMenu(menus.Menu):
    """Creates a menu for the show command"""

    def __init__(self, user, card, *args, **kwargs):
        super().__init__(timeout=30.0, delete_message_after=True, *args, **kwargs)
        self.card = card  # creates a current sort variable to sort through them
        self.user = user
        self.inventory = get_inventory(user.id)
        rates = {}
        for setting in CONVERT:
            rates[setting.split(":")[0]] = int(setting.split(":")[1])
        self.value = rates[card.rarity]  # Reads it's rarity.
        self.confirmation = None

    async def send_initial_message(self, ctx, channel):
        self.confirmation = await channel.send(
            embed=discord.Embed(title=f"Are you sure you want to convert {self.card.name} into {self.value} coin(s)?\
                                                            \nReact with ✔ to proceed.", colour=COLOR))
        return await channel.send(embed=self.card.generate_embed())

    @menus.button('📊')
    async def show_stats(self, payload):
        if is_stat_embed(self.message.embeds[0]):  # if the embed is a stats embed
            await self.message.edit(embed=self.card.generate_embed())
        # if the embed is a regular embed
        else:
            await self.message.edit(embed=self.card.generate_stats_embed())

    @menus.button('❌', position=menus.Last(2))
    async def cancel(self, payload):
        await self.message.delete()
        await self.confirmation.delete()
        return await self.message.channel.send(
            embed=discord.Embed(title=f"Conversion canceled.", colour=COLOR))
        self.stop()

    @menus.button('✔️', position=menus.Last(3))
    async def confirm(self, payload):
        remove_card(self.user, str(self.card.card_id))
        await self.message.delete()
        await self.confirmation.delete()
        await self.message.channel.send(embed=discord.Embed(
            title=f"Successfully turned {self.card.name} into {self.value} coin(s)! You now have {self.inventory['coins'] + self.value} coins.",
            colour=COLOR))
        with open("UtilityFiles/user_info.json", "r+") as f:
            stored_info = json.load(f)
            print("AHOY MATEY, 'TIS ME, THE TESTING PRINT! ARR!")
            stored_info[str(self.user.id)]["coins"] += self.value
            stored_info[str(self.user.id)]["conversions"] += 1
            dump_data(stored_info, f)
        self.stop()


class TradeConfirmMenu(menus.Menu):

    def __init__(self, user: discord.Member, card1, card2, *args, **kwargs):
        super().__init__(timeout=30.0, delete_message_after=True, *args, **kwargs)
        self.oguser = user
        self.card1 = card1
        self.card2 = card2

    async def send_initial_message(self, ctx, channel):
        return await channel.send(embed=generate_embed_trade(self))

    @menus.button('❌')
    async def cancel(self, payload):
        await self.message.delete()
        return await self.message.channel.send(
            embed=discord.Embed(title=f"Trade rejected.", colour=COLOR))
        self.stop()

    @menus.button('✔️', position=menus.Last(2))
    async def confirm(self, payload):
        remove_card(self.ctx.author, str(self.card2.card_id))
        add_card(self.ctx.author, str(self.card1.card_id))
        remove_card(self.oguser, str(self.card1.card_id))
        add_card(self.oguser, str(self.card2.card_id))

        await self.message.delete()
        await self.message.channel.send(embed=discord.Embed(
            title=f"Successfully traded {self.card1.name} and {self.card2.name}!",
            colour=COLOR))
        self.stop()


def generate_embed_trade(self):
    """Generates an embed with basic info about a card"""
    card_embed = discord.Embed(
        title=f"**Trade Request!**\nReact with ✔ to accept.\n{self.oguser.display_name}'s card: #{self.card1.card_id} | {self.card1.name}\n{self.ctx.author.display_name}'s card: #{self.card2.card_id} | {self.card2.name}",
        colour=COLOR
    )
    card_embed.set_image(url=self.card1.image_url)
    card_embed.set_thumbnail(url=self.card2.image_url)
    card_embed.add_field(name=f"**{self.card1.name}'s rarity:**", value=RARITY_MAP[self.card1.rarity])
    card_embed.add_field(name=f"**{self.card2.name}'s rarity:**", value=RARITY_MAP[self.card2.rarity])

    return card_embed


class Sorts:  # note: each function's docstring is used to show the current sort
    """Class that organizes functions used for sorting the gacha collection"""

    @staticmethod
    def id_sort(cards):
        """Number"""  # sorts by card id number
        return sorted(cards, key=lambda x: x.card_id)

    @staticmethod
    def name_sort(cards):
        """Name"""  # sorts by gacha name, alphabetically
        return sorted(cards, key=lambda x: x.name)

    @staticmethod
    def rarity_sort(cards):
        """Rarity"""  # sorts by rarity first (C, R, SR, UR), then id
        rarity_dict = {"C": [], "R": [], "SR": [], "UR": []}

        for card in cards:
            rarity_dict[card.rarity].append(card)

        sorted_cards = []
        for card_lists in rarity_dict.values():
            card_lists.sort(key=lambda x: x.card_id)
            sorted_cards += card_lists
        return sorted_cards

    @staticmethod
    def show_sort(cards):
        """Show"""  # sorts by show (alphabetically), then id
        show_dict = {}
        for card in cards:
            if card.show in show_dict.keys():
                show_dict[card.show].append(card)
            else:
                show_dict[card.show] = [card]
        # sorts the dictionary
        show_dict = {show: card_lists for show, card_lists in sorted(show_dict.items())}

        sorted_cards = []
        for card_lists in show_dict.values():
            card_lists.sort(key=lambda x: x.card_id)
            sorted_cards += card_lists
        return sorted_cards

    @staticmethod
    def amount_sort(cards):
        """Amount"""  # sorts by amount of cards held
        return sorted(cards, key=lambda x: x.amount)


SORT_LIST = [Sorts.id_sort, Sorts.name_sort, Sorts.rarity_sort, Sorts.show_sort, Sorts.amount_sort]
